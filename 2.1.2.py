from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import csv


currency_converter = {
    "GEL": 21.74, "KGS": 0.76,
    "UZS": 0.0055, "AZN": 35.68,
    "UAH": 1.64, "BYR": 23.91,
    "KZT": 0.13, "RUR": 1,
    "EUR": 59.90, "USD": 60.66
}


class Main:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.job_name = input('Введите название профессии: ')

        data_frame = Data(self.file_name, self.job_name)

        wages_by_year, jobs_number_by_year, chosen_job_wages_by_year,\
        number_of_chosen_jobs_by_year, wages_by_location,\
        location_by_number_of_jobs = data_frame.calculate_statistics()

        data_frame.write_calculations(wages_by_year, jobs_number_by_year,
                        chosen_job_wages_by_year, number_of_chosen_jobs_by_year,
                        wages_by_location, location_by_number_of_jobs)

        report = Report(self.job_name, wages_by_year, jobs_number_by_year,
                        chosen_job_wages_by_year, number_of_chosen_jobs_by_year,
                        wages_by_location, location_by_number_of_jobs)

        report.generate_excel()
        report.generate_image()


class Job:
    def __init__(self, job):
        self.name = job['name']
        self.salary_currency = job['salary_currency']
        self.salary_from = int(float(job['salary_from']))
        self.salary_to = int(float(job['salary_to']))
        self.salary_average = currency_converter[self.salary_currency]\
                              * (self.salary_to + self.salary_from) / 2
        self.year = int(job['published_at'][:4])
        self.area_name = job['area_name']


class Data:
    def __init__(self, file_name, job_name):
        self.job_name = job_name
        self.file_name = file_name

    @staticmethod
    def write_calculations(wages_by_year, jobs_number_by_year, chosen_job_wages_by_year,
                           number_of_chosen_jobs_by_year, wages_by_location,
                           location_by_number_of_jobs):
        print('Динамика уровня зарплат по годам: {0}'.format(wages_by_year))
        print('Динамика количества вакансий по годам: {0}'.format(jobs_number_by_year))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'
              .format(chosen_job_wages_by_year))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'
              .format(number_of_chosen_jobs_by_year))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(wages_by_location))
        print('Доля вакансий по городам (в порядке убывания): {0}'
              .format(location_by_number_of_jobs))

    @staticmethod
    def mean_value(input_dict):
        result = {}
        for key, values in input_dict.items():
            result[key] = int(sum(values) / len(values))
        return result

    @staticmethod
    def adder(input_dict, key, input_value):
        if key in input_dict:
            input_dict[key] += input_value
        else:
            input_dict[key] = input_value

    def read_csv(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as data_frame:
            file_head = next(csv.reader(data_frame))
            for row in csv.reader(data_frame):
                if '' not in row and len(row) == len(file_head):
                    yield dict(zip(file_head, row))

    def calculate_statistics(self):
        location = {}
        jobs_counter = 0
        wage = {}
        wage_vacancy_name = {}

        jobs_counter = self.fill_dicts(jobs_counter, location, wage, wage_vacancy_name)

        jobs_amount_by_name = dict([(key, len(value)) for key, value in wage_vacancy_name.items()])
        jobs_amount = dict([(key, len(value)) for key, value in wage.items()])

        if not wage_vacancy_name:
            jobs_amount_by_name = dict([(key, 0) for key, value in jobs_amount.items()])
            wage_vacancy_name = dict([(key, [0]) for key, value in wage.items()])

        mean_wage = self.mean_value(wage)
        jobs_number_by_year = self.mean_value(wage_vacancy_name)
        chosen_job_wages_by_year = self.mean_value(location)
        number_of_chosen_jobs_by_year = {}

        for year, salaries in location.items():
            number_of_chosen_jobs_by_year[year] = round(len(salaries) / jobs_counter, 4)

        number_of_chosen_jobs_by_year = list(
            filter(lambda a:
                   a[-1] >= 0.01,
                   [(key, value) for key, value in number_of_chosen_jobs_by_year.items()]))

        number_of_chosen_jobs_by_year.sort(key=lambda b: b[-1], reverse=True)
        wages_by_location = number_of_chosen_jobs_by_year.copy()
        number_of_chosen_jobs_by_year = dict(number_of_chosen_jobs_by_year)

        chosen_job_wages_by_year = list(
            filter(lambda c:
                   c[0] in list(number_of_chosen_jobs_by_year.keys()),
                   [(key, value) for key, value in chosen_job_wages_by_year.items()]))

        chosen_job_wages_by_year.sort(key=lambda d: d[-1], reverse=True)
        chosen_job_wages_by_year = dict(chosen_job_wages_by_year[:10])
        wages_by_location = dict(wages_by_location[:10])

        return mean_wage, jobs_amount, jobs_number_by_year,\
               jobs_amount_by_name, chosen_job_wages_by_year, wages_by_location

    def fill_dicts(self, jobs_counter, location, wage, wage_vacancy_name):
        for job_dict in self.read_csv():
            job = Job(job_dict)
            self.adder(wage, job.year, [job.salary_average])
            if job.name.find(self.job_name) != -1:
                self.adder(wage_vacancy_name, job.year, [job.salary_average])
            self.adder(location, job.area_name, [job.salary_average])
            jobs_counter += 1
        return jobs_counter


class Report:
    def __init__(self, jobs_name, wages_by_year, jobs_number_by_year,
                 chosen_job_wages_by_year, number_of_chosen_jobs_by_year,
                 wages_by_location, location_by_number_of_jobs):
        self.workbook = Workbook()
        self.jobs_name = jobs_name
        self.wages_by_year = wages_by_year
        self.jobs_number_by_year = jobs_number_by_year
        self.chosen_job_wages_by_year = chosen_job_wages_by_year
        self.number_of_chosen_jobs_by_year = number_of_chosen_jobs_by_year
        self.wages_by_location = wages_by_location
        self.location_by_number_of_jobs = location_by_number_of_jobs

    def generate_excel(self):
        widths_column = []
        first_work_sheet = self.workbook.active
        first_work_sheet.title = 'Статистика по годам'
        first_work_sheet.append(['Год', 'Средняя зарплата',
            'Средняя зарплата - ' + self.jobs_name, 'Количество вакансий',
                'Количество вакансий - ' + self.jobs_name])

        for year in self.wages_by_year.keys():
            first_work_sheet.append([year, self.wages_by_year[year],
                self.chosen_job_wages_by_year[year], self.jobs_number_by_year[year],
                self.number_of_chosen_jobs_by_year[year]])

        head = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - '
                 + self.jobs_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.jobs_name]]
        self.set_column_width_first_sheet(first_work_sheet, head, widths_column)
        head = []
        head.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])

        for (first_city, first_val), (second_city, second_val) in \
                zip(self.wages_by_location.items(), self.location_by_number_of_jobs.items()):
            head.append([first_city, first_val, '', second_city, second_val])
        second_work_sheet = self.workbook.create_sheet('Статистика по городам')
        for row in head:
            second_work_sheet.append(row)
        widths_column = []
        self.set_column_width_second_sheet(head, second_work_sheet, widths_column)

        font_bold = Font(bold=True)
        for column in 'ABCDE':
            first_work_sheet[column + '1'].font = font_bold
            second_work_sheet[column + '1'].font = font_bold
        for index, _ in enumerate(self.wages_by_location):
            second_work_sheet['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')
        for row in range(len(head)):
            for column in 'ABDE':
                second_work_sheet[column + str(row + 1)].border = Border(left=thin,
                bottom=thin, right=thin, top=thin)
        self.wages_by_year[1] = 1
        for row, _ in enumerate(self.wages_by_year):
            for column in 'ABCDE':
                first_work_sheet[column + str(row + 1)]\
                    .border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.workbook.save('report.xlsx')

    def set_column_width_first_sheet(self, first_work_sheet, head, widths_column):
        for row in head:
            for i, cell in enumerate(row):
                if len(widths_column) > i:
                    if len(cell) > widths_column[i]:
                        widths_column[i] = len(cell)
                else:
                    widths_column += [len(cell)]
        for i, column_width in enumerate(widths_column, 1):
            first_work_sheet.column_dimensions[get_column_letter(i)].width = column_width + 2

    def set_column_width_second_sheet(self, head, second_work_sheet, widths_column):
        for row in head:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths_column) > i:
                    if len(cell) > widths_column[i]:
                        widths_column[i] = len(cell)
                else:
                    widths_column += [len(cell)]
        for i, column_width in enumerate(widths_column, 1):
            second_work_sheet.column_dimensions[get_column_letter(i)].width = column_width + 2

    def generate_image(self):
        fig, ((first_sub_plot, second_sub_plot),
              (third_sub_plot, fourth_sub_plot)) = plt.subplots(nrows=2, ncols=2)

        self.draw_wages_by_year(first_sub_plot)
        self.draw_jobs_number_by_year(second_sub_plot)
        self.draw_wages_by_location(third_sub_plot)
        self.draw_location_by_number_of_jobs(fourth_sub_plot)

        plt.tight_layout()
        plt.savefig('graph.png')

    def draw_wages_by_year(self, first_sub_plot):
        self.wages_by_year.pop(1)
        first_bar = first_sub_plot.bar(np.array(list(self.wages_by_year.keys())) - 0.4,
                                  self.wages_by_year.values(), width=0.4)
        second_bar = first_sub_plot.bar(np.array(list(self.wages_by_year.keys())),
                                  self.chosen_job_wages_by_year.values(), width=0.4)
        first_sub_plot.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        first_sub_plot.grid(axis='y')
        first_sub_plot.legend((first_bar[0], second_bar[0]), ('средняя з/п', 'з/п '
                                                   + self.jobs_name.lower()), prop={'size': 8})
        first_sub_plot.set_xticks(np.array(list(self.wages_by_year.keys())) - 0.2,
                                  list(self.wages_by_year.keys()), rotation=90)
        first_sub_plot.xaxis.set_tick_params(labelsize=8)
        first_sub_plot.yaxis.set_tick_params(labelsize=8)

    def draw_jobs_number_by_year(self, second_sub_plot):
        second_sub_plot.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        first_bar = second_sub_plot.bar(np.array(list(self.jobs_number_by_year.keys())) - 0.4,
                                   self.jobs_number_by_year.values(), width=0.4)
        second_bar = second_sub_plot.bar(np.array(list(self.jobs_number_by_year.keys())),
                                   self.number_of_chosen_jobs_by_year.values(), width=0.4)
        second_sub_plot.legend((first_bar[0], second_bar[0]), ('Количество вакансий',
                                                               'Количество вакансий\n'
                                                    + self.jobs_name.lower()), prop={'size': 8})
        second_sub_plot.set_xticks(np.array(list(self.jobs_number_by_year.keys())) - 0.2,
                                   list(self.jobs_number_by_year.keys()), rotation=90)
        second_sub_plot.grid(axis='y')
        second_sub_plot.xaxis.set_tick_params(labelsize=8)
        second_sub_plot.yaxis.set_tick_params(labelsize=8)

    def draw_wages_by_location(self, third_sub_plot):
        third_sub_plot.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        third_sub_plot.barh(list([str(a).replace(' ', '\n').replace('-', '-\n')
                                  for a in reversed(list(self.wages_by_location.keys()))]),
                            list(reversed(list(self.wages_by_location.values()))),
                            color='blue', height=0.5, align='center')
        third_sub_plot.yaxis.set_tick_params(labelsize=6)
        third_sub_plot.xaxis.set_tick_params(labelsize=8)
        third_sub_plot.grid(axis='x')

    def draw_location_by_number_of_jobs(self, fourth_sub_plot):
        fourth_sub_plot.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other_locations = 1 - sum([value for value in self.location_by_number_of_jobs.values()])
        fourth_sub_plot.pie(list(self.location_by_number_of_jobs.values()) + [other_locations],
                            labels=list(self.location_by_number_of_jobs.keys()) + ['Другие'],
                            textprops={'fontsize': 6})


if __name__ == '__main__':
    Main()